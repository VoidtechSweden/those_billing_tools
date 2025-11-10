import openpyxl

from billing.fields.billable_item_field import BillableItemField
from billing.fields.invoice_field import InvoiceField
from billing.fields.normal_hours_field import NormalHoursField
from utils import input_tools

ITEM_DESCRIPTION_COLUMN = "B"
ITEM_VALUE_COLUMN = "C"
ITEM_UNIT_COLUMN = "D"


class ItemFieldsGenerator:
    """
    A class responsible for generating fields for billable items for an invoice, based on the item fields in the provided template file.
    """

    def __init__(self, template_file: str) -> None:
        self._template_file = template_file

    def generate_item_fields(self) -> list[InvoiceField]:
        """
        Generate item fields for the invoice based on the rows of billable items in the template file
        """

        item_fields: list[InvoiceField] = []
        optional_fields: list[tuple[str, str]] = []

        # Normal hours are always mandatory
        normal_hours = NormalHoursField()
        item_fields.append(normal_hours)

        wb = openpyxl.load_workbook(self._template_file, read_only=True)
        assert wb is not None, "Workbook could not be loaded"
        ws = wb.active
        assert ws is not None, "Worksheet could not be loaded"

        normal_hours_field = normal_hours.get_field()
        item_start_row = int(normal_hours_field[1:])

        # Find all optional billable items in the template
        for i in range(1, 20):
            item_row = item_start_row + i

            item_description = ws[f"{ITEM_DESCRIPTION_COLUMN}{item_row}"].value
            item_unit = ws[f"{ITEM_UNIT_COLUMN}{item_row}"].value

            if item_description is None or item_unit is None:
                # If either description or unit is missing, skip this row
                continue
            optional_fields.append(
                (f"{ITEM_VALUE_COLUMN}{item_row}", f"{item_description} ({item_unit})")
            )

        # Let user select optional items to include
        while len(optional_fields) > 0:
            descriptions = [field[1] for field in optional_fields]
            selected, selected_index = input_tools.select_indexed_item(
                "Select an optional billing item to include",
                descriptions,
                optional_selection=True,
            )
            if selected_index is None:
                break

            selected_field = BillableItemField(
                optional_fields[selected_index - 1][0],
                optional_fields[selected_index - 1][1],
            )
            item_fields.append(selected_field)
            optional_fields.pop(selected_index - 1)

        return item_fields
