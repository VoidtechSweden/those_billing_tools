from datetime import datetime as DateTime
import os
from openpyxl import load_workbook

from utils import basic_tools, input_tools

from billing import billing_tools

from config.config import Configuration

if __name__ == "__main__":

    print(f"Creating invoice for {Configuration.get('billing', 'company')}")

    ### Get invoice date ###
    date = input_tools.input_date("Enter date (YYYY-MM-DD)", DateTime.now())
    print("Billing period: " + billing_tools.month_to_shortstring(date.month))

    ### Get billed hours ###
    hours = input_tools.input_number("Enter number of worked hours")

    ### Get new invoice number ###
    invoice_number = input_tools.input_number(
        f"Enter invoice number", billing_tools.get_latest_invoice_nr() + 1
    )
    if billing_tools.invoice_already_exists(invoice_number):
        basic_tools.paused_exit(
            f"Invoice with number {invoice_number} already exists",
            "Exiting to avoid overwriting an existing file.",
        )

    ### Generate invoice name ###
    invoice_path = billing_tools.create_invoice_path(invoice_number)
    print(f"File will be saved as: '{invoice_path}'")

    ### Get the correct template to use ###
    template_prefix = Configuration.get('billing', 'template_prefix')
    template_path = Configuration.get('billing', 'template_path')
    template_files = [
        os.path.join(template_path, f)
        for f in os.listdir(template_path)
        if os.path.isfile(os.path.join(template_path, f))
        and f.startswith(template_prefix)
        and f.endswith(".xlsx")
    ]
    if not template_files:
        basic_tools.paused_exit(
            f"No template file with prefix '{template_prefix}' found in '{template_path}'"
        )
    elif len(template_files) == 1:
        template_file = template_files[0]
    else:
        template_file = input_tools.select_indexed_item(
            "Select a billing template", template_files
        )
    print(f"Using billing template: '{template_file}'")

    ### Create and save the spreadsheet bill ###
    wb = load_workbook(template_file)
    ws = wb.active
    ws[billing_tools.BillFields.NUMBER_FIELD] = invoice_number
    ws[billing_tools.BillFields.MONTH_FIELD] = billing_tools.month_to_shortstring(
        date.month
    )
    ws[billing_tools.BillFields.DATE_FIELD] = date.strftime("%Y-%m-%d")
    ws[billing_tools.BillFields.HOURS_FIELD] = hours
    wb.save(invoice_path)

    # Try to convert to PDF
    pdf_path = billing_tools.convert_to_pdf(invoice_path)

    print("\nSummary")
    print("================================")
    print(f"Template: '{os.path.relpath(template_file)}'")
    print(f"Number: {invoice_number}")
    print(f"Period: {billing_tools.month_to_shortstring(date.month)}")
    print(f"Date: {date.strftime('%Y-%m-%d')}")
    print(f"Hours: {hours} hours")
    print(f"XLSX: {os.path.relpath(invoice_path)}")
    print(f"PDF: {os.path.relpath(pdf_path) if pdf_path else 'NO'}")
    print("================================")
    basic_tools.paused_exit("Invoice created")
